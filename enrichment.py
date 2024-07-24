import pandas as pd
import customtkinter as ctk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import io
import json
import os

# Define global variables for storing SKU and SEQ/NAME data
sku_data_list = []
seq_name_data_list = []
taxonomy_groups = {}
skip_taxonomies = set()  # To keep track of skipped taxonomies

# File to store past runs
history_file = 'enrichment_history.json'

# Initialize the main window
root = ctk.CTk()
root.title("SKU Data Processor")
root.geometry("600x800")

ctk.set_appearance_mode("dark")
# Custom Messagebox class
class CTkMessagebox(ctk.CTkToplevel):
    def __init__(self, master=None, title="Message", message=""):
        super().__init__(master)
        self.title(title)
        self.geometry("300x150")

        self.label = ctk.CTkLabel(self, text=message)
        self.label.pack(pady=20)

        self.button = ctk.CTkButton(self, text="OK", command=self.destroy)
        self.button.pack(pady=10)

# Function to show warning messagebox
def show_warning(title, message):
    messagebox = CTkMessagebox(title=title, message=message)
    messagebox.grab_set()  # Make the messagebox modal

# Function to show error messagebox
def show_error(title, message):
    messagebox = CTkMessagebox(title=title, message=message)
    messagebox.grab_set()  # Make the messagebox modal

# Function to show info messagebox
def show_info(title, message):
    messagebox = CTkMessagebox(title=title, message=message)
    messagebox.grab_set()  # Make the messagebox modal

def process_sku_data(sku_data):
    try:
        # Read SKU data
        sku_df = pd.read_csv(io.StringIO(sku_data), delimiter='\t')
        print(f"SKU DataFrame:\n{sku_df}")  # Debugging output

        # Check if 'Purpose' column contains 'Packaging' and overwrite 'Attribute Value' column data with 'x'
        if 'Purpose' in sku_df.columns and 'Attribute Value' in sku_df.columns:
            sku_df.loc[sku_df['Purpose'].str.contains('Packaging', case=False, na=False), 'Attribute Value'] = 'x'

        return sku_df
    except Exception as e:
        show_error("Error", f"Error processing SKU data: {str(e)}")
        return None
    
    #defines clear data
def clear_data():
    global sku_data_list, seq_name_data_list, taxonomy_groups, skip_taxonomies
    sku_data_list = []
    seq_name_data_list = []
    taxonomy_groups = {}
    skip_taxonomies = set()

    text_area.delete("1.0", ctk.END)
    sku_listbox.delete("1.0", ctk.END)
    sku_listbox.configure(state="disabled")

    step_2_button.config(state=ctk.DISABLED)

    add_another_sku_button.config(state=ctk.DISABLED)
    complete_button.config(state=ctk.DISABLED)
    skip_button.config(state=ctk.DISABLED)

    show_info("Info", "All data cleared. Ready to start fresh.")

def save_history():
    history = {
        "runs": []
    }
    if os.path.exists(history_file):
        with open(history_file, 'r') as file:
            history = json.load(file)
    
    run_data = {
        "sku_data": sku_data_list,
        "seq_name_data": seq_name_data_list
    }
    
    history["runs"].insert(0, run_data)
    history["runs"] = history["runs"][:3]  # Keep only last 3 runs
    
    with open(history_file, 'w') as file:
        json.dump(history, file)

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def create_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Enrichment"
        
        # Add SKU headers
        sku_headers = [
            'SKU #', 'SKU Title (English)', 'Attribute value of PKG Product Identity Modifier (English, DEFAULT)', 
            'Attribute value of PKG Product Identity (English, DEFAULT)', 'Attribute value of PKG Custom Callout (English, DEFAULT)', 
            'Attribute value of Ideal for (English, DEFAULT)', 'Structure group(s) (Brands Structure)', 'PKG Current Package Type', 
            'Structure assignments (Selling Taxonomy)', 'Structure assignments (Promotion Structure)', 'MMS Item Status', 'Vendor'
        ]
        ws.append(sku_headers)
        
        # Highlight and bold the header row
        fill_red = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = fill_red
            cell.font = bold_font
        
        # Add SKU data rows starting at row 2
        for sku_data in sku_data_list:
            sku_df = pd.read_csv(io.StringIO(sku_data), delimiter='\t')
            for _, row in sku_df.iterrows():
                ws.append(row.tolist())
                # Extract Structure assignments and initialize combined data
                taxonomy = row['Structure assignments (Selling Taxonomy)']
                if taxonomy not in taxonomy_groups:
                    taxonomy_groups[taxonomy] = []
        
        # Add secondary headers
        ws.append([])
        secondary_headers = ['Name (English)', 'Attribute value (English, DEFAULT)', 'Purpose']
        ws.append([""] + secondary_headers)
        
        # Add green row after secondary headers
        fill_light_green = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill_light_green
        
        # Write SEQ and NAME data to Excel directly after each Step 2
        for seq_name_data in seq_name_data_list:
            seq_name_df = pd.read_csv(io.StringIO(seq_name_data), delimiter='\t')
            for _, row in seq_name_df.iterrows():
                name = row.get('Name (English)', '')
                attribute_value = row.get('Attribute value (English, DEFAULT)', '')
                purpose = row.get('Purpose', '')
                ws.append([""] + [name, attribute_value, purpose])
            # Add green row after each set of SEQ and NAME data
            ws.append([])  # Insert empty row
            for cell in ws[ws.max_row]:
                cell.fill = fill_light_green

        # Auto-adjust column widths
        auto_adjust_column_width(ws)
        
        # Save the workbook
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filename:
            wb.save(filename)
            save_history()
            show_info("Success", f"File saved as {filename}")
    except Exception as e:
        show_error("Error", f"An error occurred while creating the Excel file: {str(e)}")

def paste_sku_data():
    try:
        sku_data = text_area.get("1.0", "end-1c").strip()  # Use "end-1c" to exclude the trailing newline character
        if not sku_data:
            raise ValueError("SKU data is empty.")
        
        sku_df = process_sku_data(sku_data)
        if sku_df is not None:
            sku_data_list.append(sku_df.to_csv(sep='\t', index=False))
            text_area.delete("1.0", "end")

            # Extract SKU # and insert into the top of the listbox
            try:
                sku_number = sku_df['SKU #'].iloc[0]
                sku_listbox.configure(state="normal")
                sku_listbox.insert("1.0", sku_number + "\n")  # Insert SKU # into listbox
                sku_listbox.configure(state="disabled")
            except Exception as e:
                show_error("Error", f"Error inserting SKU # into listbox: {str(e)}")

            step_2_button.config(state=ctk.NORMAL)
            
            show_info("Success", "Step 1 completed. Now proceed to Step 2.")
    except Exception as e:
        show_error("Error", f"Step 1 Error: {str(e)}")



def paste_seq_name_data():
    try:
        seq_name_data = text_area.get("1.0", ctk.END).strip()
        if not seq_name_data:
            raise ValueError("SEQ and NAME data is empty.")
        seq_name_data_list.append(seq_name_data)
        text_area.delete("1.0", ctk.END)
        add_another_sku_button.configure(state=ctk.NORMAL)
        complete_button.configure(state=ctk.NORMAL)
        skip_button.configure(state=ctk.NORMAL)  # Enable skip button
        show_info("Success", "Step 2 completed. You can add another SKU, skip this data, or complete the process.")
    except Exception as e:
        show_error("Error", f"Step 2 Error: {str(e)}")

def add_another_sku():
    step_2_button.configure(state=ctk.DISABLED)
    show_info("Info", "Ready for another SKU. Please paste SKU data for the next product.")

def skip_step_2():
    global skip_taxonomies
    # Optionally, you can prompt the user for which taxonomy to skip
    skip_taxonomies = {taxonomy for taxonomy in taxonomy_groups.keys()}
    show_info("Info", "All data for Step 2 has been skipped for the current SKU.")

def complete_enrichment():
    create_excel()

def populate_sku_listbox():
    try:
        with open(history_file, 'r') as file:
            data = json.load(file)
            for run in data.get("runs", []):
                for sku in run.get("sku_data", []):
                    sku_df = pd.read_csv(io.StringIO(sku), delimiter='\t')
                    for _, row in sku_df.iterrows():
                        sku_listbox.configure(state="normal")
                        sku_listbox.insert(1.0, row['SKU #'])
                        sku_listbox.configure(state="disabled")
    except FileNotFoundError:
        pass  # No history file exists yet
    except json.JSONDecodeError:
        show_error("Error", "Failed to decode JSON file.")
    except Exception as e:
        show_error("Error", f"Failed to populate SKU listbox: {str(e)}")

        # Create a label as a title for the input area
title_label = ctk.CTkLabel(root, text="Input Data Here",)  # You can adjust the font and size as needed
title_label.pack(pady=(20, 0))  # Adding some padding above the label

# Initialize the Text area for input
text_area = ctk.CTkTextbox(root, width=400, height=200, fg_color="#181818", corner_radius=10)
text_area.pack(pady=(10, 20))  # Adding padding to separate from the title and below the textbox

title_label = ctk.CTkLabel(root, text="Previous SKU's",)  # You can adjust the font and size as needed
title_label.pack(pady=(20, 0))  # Adding some padding above the label
# Initialize the SKU Listbox
sku_listbox = ctk.CTkTextbox(root)
sku_listbox.pack(pady=10)
sku_listbox.configure(state="disabled")

# Populate SKU listbox with past data
populate_sku_listbox()



# Initialize the buttons
paste_sku_button = ctk.CTkButton(root, text="Process SKU Data (Step 1)", fg_color="#0057B8", command=paste_sku_data)
paste_sku_button.pack(pady=10)

step_2_button = ctk.CTkButton(root, text="Process SEQ and NAME Data (Step 2)", fg_color="#0057B8", state=ctk.DISABLED, command=paste_seq_name_data)
step_2_button.pack(pady=10)

skip_button = ctk.CTkButton(root, text="Skip Step 2 (Same Class Procedure)", fg_color="#562C2C", state=ctk.DISABLED, command=skip_step_2)
skip_button.pack(pady=10)

add_another_sku_button = ctk.CTkButton(root, text="Add Another SKU", fg_color="#FFD100", state=ctk.DISABLED, command=add_another_sku)
add_another_sku_button.pack(pady=10)

complete_button = ctk.CTkButton(root, text="Complete Enrichment", fg_color="#37696C", state=ctk.DISABLED, command=complete_enrichment)
complete_button.pack(pady=10)

clear_button = ctk.CTkButton(root, text="Clear Data", fg_color="#D7263D", command=clear_data)
clear_button.pack(pady=10)


# Run the main loop
root.mainloop()
